import openpyxl
from openpyxl.styles import Alignment

# Definindo as informações sobre os meses
meses = [
    {'nome': 'Janeiro', 'dias': 31},
    {'nome': 'Fevereiro', 'dias': 28},
    {'nome': 'Março', 'dias': 31},
    {'nome': 'Abril', 'dias': 30},
    {'nome': 'Maio', 'dias': 31},
    {'nome': 'Junho', 'dias': 30},
    {'nome': 'Julho', 'dias': 31},
    {'nome': 'Agosto', 'dias': 31},
    {'nome': 'Setembro', 'dias': 30},
    {'nome': 'Outubro', 'dias': 31},
    {'nome': 'Novembro', 'dias': 30},
    {'nome': 'Dezembro', 'dias': 31}
]

# Criando o arquivo Excel
wb = openpyxl.Workbook()

# Iterando pelos meses
for mes in meses:
    # Criando uma nova planilha para o mês atual
    ws = wb.create_sheet(mes['nome'])

    # Definindo os dias da semana
    dias_semana = ['Domingo', 'Segunda-feira', 'Terça-feira', 'Quarta-feira',
                   'Quinta-feira', 'Sexta-feira', 'Sábado']

    # Definindo a data inicial (ano começou em um domingo)
    data_atual = 1 if mes['nome'] == 'Janeiro' else 0
    dia_semana_atual = 0 if mes['nome'] == 'Janeiro' else 1

    # Preenchendo os cabeçalhos das colunas com os dias da semana
    for col, dia_semana in enumerate(dias_semana, start=1):
        ws.cell(row=1, column=col, value=dia_semana)

    # Preenchendo as datas e deixando um espaço abaixo para aulas
    for row in range(2, mes['dias'] + 2):
        # Preenchendo a data
        ws.cell(row=row, column=1, value=data_atual)

        # Centralizando o conteúdo na célula
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center', vertical='center')

        # Preenchendo o espaço abaixo da data para aulas
        ws.cell(row=row + 1, column=1, value='')

        # Atualizando a data e o dia da semana
        data_atual += 1
        dia_semana_atual += 1

        # Verificando se é necessário reiniciar o dia da semana no próximo mês
        if dia_semana_atual == 7:
            dia_semana_atual = 0

    # Ajustando a largura das colunas para caber o conteúdo
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

# Removendo a planilha inicial padrão
wb.remove(wb['Sheet'])

# Salvando o arquivo Excel
wb.save('plano_de_aula.xlsx')
